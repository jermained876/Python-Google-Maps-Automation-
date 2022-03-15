from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('locations.xlsx')
ws = wb['Sheet1']


print(ws.max_row)
print(ws.max_column)



for row in range(2,ws.max_row):
    print(ws['A' + str(row)].value)
    print(ws['B' + str(row)].value)