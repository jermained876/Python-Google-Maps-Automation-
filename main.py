import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


driver = webdriver.Chrome('chromedriver.exe')
driver.get("https://www.google.com/maps")
wb = load_workbook('locations.xlsx')
ws = wb['Sheet1']


driver.implicitly_wait(8)

directionButton = driver.find_element(By.ID, "xoLGzf-T3iPGc")
directionButton.click()
rows = ws.max_row+1

for row in range(2,rows):
    print(ws['A' + str(row)].value)
    print(ws['B' + str(row)].value)

    from_textbox = driver.find_element(By.XPATH, "//*[@id='sb_ifc51']/input")
    to_textbox = driver.find_element(By.XPATH, "//*[@id='sb_ifc52']/input")
    from_textbox.clear()
    to_textbox.clear()

    from_textbox.send_keys(ws['A' + str(row)].value)
    to_textbox.send_keys(ws['B' + str(row)].value)

    to_textbox.send_keys(Keys.ENTER)

    try:
        distance_km = driver.find_element(By.XPATH, "//div[contains(text(),'km')]")
    except:
        distance_km = driver.find_element(By.XPATH, "//div[contains(text(),'miles')]")

    try:
        distance_car = driver.find_element(By.XPATH, "//div[contains(text(),'min')]")
    except:
        distance_car = driver.find_element(By.XPATH, "//div[contains(text(),'hr')]")

    ws['C' + str(row)] = distance_km.text
    ws['D' + str(row)] = distance_car.text
    print(distance_km.text)
    print(distance_car.text)
    time.sleep(3)

wb.save('locations.xlsx')

driver.close()